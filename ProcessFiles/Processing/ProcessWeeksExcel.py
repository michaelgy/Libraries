from BaseClasses import ProcessFileWF
from openpyxl import load_workbook
import re

class main(ProcessFileWF):
    def __init__(self, input_file, output_file):
        super().__init__(input_file, output_file)
        self.process()
    
    def process(self):
        wbinput = load_workbook(filename = self.input_file)
        pattern = ".*-.*-.*-.*"

        wboutput = load_workbook(filename = self.output_file)
        wsout = wboutput["Reportes"]
        lastrow = len(wsout["B"]) # last row with data
        pass

if __name__ == "__main__":
    """Testing"""
    output_file="Z:\\Informacion de estudiante en practica\\Archivos para indicadores de mantto\\MTTO CORRECTIVO - DMS\Anual\\Datos2020S01-08.xlsm"
    input_file="Z:\\Informacion de estudiante en practica\\Archivos para indicadores de mantto\\MTTO CORRECTIVO - DMS\\2020\\SEMANAS\\Semana 01\\Mtto Correctivo Semana 01.xlsx"
    main(input_file, output_file)
    