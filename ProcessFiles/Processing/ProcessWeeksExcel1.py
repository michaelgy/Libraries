"""
Funciona para las semanas 01 a 08 de los reportes del 2020
"""
from .BaseClasses import ProcessFileWF
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re
from openpyxl.styles.colors import Color
from copy import copy
from pathlib import Path

class main(ProcessFileWF):
    def __init__(self, input_file:str, output_file:str):
        super().__init__(input_file, output_file)
        self.process()
    
    def process_sheet(self, input_sheet:"Worksheet object", output_sheet:"Worksheet object", output_startrow:int):
        """input_sheet: hoja de datos donde se van a extraer los datos
        output_sheet: hoja de datos donde se van a guardar los datos
        output_startrow: fila de output_sheet donde se comienzan a escribir los datos
        """
        print(input_sheet.title)

        rd = 0
        for c in input_sheet["B"]:
            if c.value == "EQUIPO" or c.value == "EQUIPOS":
                rd = c.row
                break
        
        t = 0
        if rd:
            rdi =rd+1
            columna_busqueda = 2
            columna_inicio_datos = 2 #Columna de input_sheet donde comienzan los datos
            columna_final_datos = 8 #Columna de input_sheet donde terminan los datos
            columna_inicio = 2 #Columna de output_sheet donde comienzan los datos
            v = input_sheet.cell(row=rdi, column=columna_busqueda).value.strip().lower()
            if not re.fullmatch(".*sin.*novedad.*", v): #Verificar que hayan novedades para copiar
                rdf =rd+1

                v = input_sheet.cell(row=rdf, column=columna_busqueda)
                while(v.value and len(v.value.strip())>=4):
                    rdf+=1
                    v = input_sheet.cell(row=rdf, column=columna_busqueda)
                for r in range(rdi,rdf):
                    for c in range(columna_inicio_datos, columna_final_datos+1):
                        cout = output_sheet.cell(row=output_startrow+r-rdi, column=columna_inicio+c-columna_inicio_datos)
                        cin = input_sheet.cell(row=r, column=c)
                        cout.value = cin.value
                        """
                        if r == rdi+1:
                            print("-----"*20)
                            print(cin.fill.bgColor)
                            print("-----"*20)
                        """
                        cout.fill =  copy(cin.fill)
                    output_sheet.cell(row=output_startrow+r-rdi, column=columna_inicio+columna_final_datos-columna_inicio_datos+1).value = Path(self.input_file).name

                    t = t +1
        print("\t {} {} {}".format("Se escribieron", t, "filas"))
    def process(self):
        wbinput = load_workbook(filename = self.input_file, keep_vba=True)
        pattern = ".*-.*-.*"
        fm = lambda name: re.fullmatch(pattern, name)

        wboutput = load_workbook(filename = self.output_file,keep_vba=True)
        wsout = wboutput["Reportes"]

        for n in filter(fm, wbinput.sheetnames):
            lastrow = len(wsout["B"]) # last row with data
            self.process_sheet(wbinput[n], wsout, lastrow+1)
        
        fname, extension = self.output_file.split(".")
        wboutput.save(self.output_file)

if __name__ == "__main__":
    """Testing"""
    output_file="Z:\\Informacion de estudiante en practica\\Archivos para indicadores de mantto\\MTTO CORRECTIVO - DMS\Anual\\2020S01-08.xlsm"
    input_file="Z:\\Informacion de estudiante en practica\\Archivos para indicadores de mantto\\MTTO CORRECTIVO - DMS\\2020\\SEMANAS\\Semana 08\\Mtto Correctivo Semana 08.xlsx"
    main(input_file, output_file)