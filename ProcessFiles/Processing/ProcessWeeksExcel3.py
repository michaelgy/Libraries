"""
Funciona para las semanas 21 a 29 de los reportes del 2020
"""
from .BaseClasses import ProcessFileWF
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re
from openpyxl.styles.colors import Color
from copy import copy
from pathlib import Path

class main(ProcessFileWF):
    def __init__(self, input_file:str, output_file:str,write_to_file=True):
        super().__init__(input_file, output_file)
        self.process(write_to_file=write_to_file)
    
    def cell(self, input_sheet:"Worksheet object", value,regex_flag = 0, init_row:int=1, max_row:int = None, init_column=1, max_column:int = 30)-> list:
        """input_sheet: hoja de datos donde se van a extraer los datos
        value: valor esperado en la celda, si es un string se puede utilizar regex

        return: a list with tuples with rows and columns
        """
        pos = []
        if isinstance(value,str):
            for r in range(init_row,max_row+1 if max_row else input_sheet.max_row+1):
                for c in range(init_column,max_column):
                    v = input_sheet.cell(row=r, column=c).value
                    if isinstance(v,str) and re.fullmatch(value, v, regex_flag):
                        pos.append((r,c))
        else:
            for r in range(1,max_row+1 if max_row else input_sheet.max_row+1):
                for c in range(1,max_column):
                    if value == input_sheet.cell(row=r, column=c):
                        pos.append((r,c))
        return pos
    
    def process_sheet(self, input_sheet:"Worksheet object", output_sheet:"Worksheet object", output_startrow:int, output_startrow_micro:int):
        """input_sheet: hoja de datos donde se van a extraer los datos
        output_sheet: hoja de datos donde se van a guardar los datos
        output_startrow: fila de output_sheet donde se comienzan a escribir los datos
        """
        print(input_sheet.title)

        rd = 0
        t = 0
        for c in input_sheet["B"]:
            if c.value == "EQUIPO":
                t += 1
            if t==2:
                rd = c.row
                break
        
        t = 0
        if rd:
            rdi = rd + 1
            rdf = rd + 1
            columna_busqueda = 2
            columna_inicio_datos = 2 #Columna de input_sheet donde comienzan los datos
            columna_final_datos = 13 #Columna de input_sheet donde terminan los datos
            fila_inicio_micro, columna_inicio_micro = self.cell(input_sheet, "MICROPARADAS")[0] #Fila y Columna de input_sheet donde comienzan los datos de microparadas
            fila_inicio_micro += 2
            columna_final_micro = columna_inicio_micro+2 #Columna de input_sheet donde terminan los datos de microparadas
            columna_inicio = 2 #Columna de output_sheet donde comienzan los datos
            columna_inicio_m = 16 #Columna de output_sheet donde comienzan los datos de microparadas
            v = input_sheet.cell(row=rdi, column=columna_busqueda).value
            if isinstance(v,str) and not re.fullmatch(".*sin.*novedad.*", v): #Verificar que hayan novedades para copiar
                v = input_sheet.cell(row=rdf, column=columna_busqueda)
                while(v.value and len(v.value.strip())>=4):
                    rdf+=1
                    v = input_sheet.cell(row=rdf, column=columna_busqueda)
                for r in range(rdi,rdf):
                    for c in range(columna_inicio_datos, columna_final_datos+1):
                        cout = output_sheet.cell(row=output_startrow+r-rdi, column=columna_inicio+c-columna_inicio_datos)
                        cin = input_sheet.cell(row=r, column=c)
                        cout.value = cin.value
                        cout.fill =  copy(cin.fill)
                    output_sheet.cell(row=output_startrow+r-rdi, column=columna_inicio+columna_final_datos-columna_inicio_datos+1).value = Path(self.input_file).name
                    t = t +1
            
            v = input_sheet.cell(row=fila_inicio_micro, column=columna_inicio_micro).value
            v = False if not v else not re.fullmatch(".*Total.*",v) if isinstance(v,str) else True
            k = 0
            while v:
                for c in range(columna_inicio_micro, columna_final_micro+1):
                    cin = input_sheet.cell(row=fila_inicio_micro+k, column=c)
                    cout = output_sheet.cell(row=output_startrow_micro+k, column=columna_inicio_m+c-columna_inicio_micro)
                    cout.value = cin.value
                k+=1
                v = input_sheet.cell(row=fila_inicio_micro+k, column=columna_inicio_micro).value
                v = False if not v else not re.fullmatch(".*total.*",v,re.I) if isinstance(v,str) else True

        print("\t {} {} {}".format("Se escribieron", t, "filas"))

    def process(self,write_to_file=True):
        wbinput = load_workbook(filename = self.input_file, keep_vba=True)
        pattern = ".*SEMANA.*(\d|\d\d).*"
        fm = lambda name: re.fullmatch(pattern, name)

        wboutput = load_workbook(filename = self.output_file,keep_vba=True)
        wsout = wboutput["Reportes"]

        for n in filter(fm, wbinput.sheetnames):
            lastrow = len(wsout["B"]) # last row with data

            lastrowm = 1
            while lastrowm <= lastrow and wsout.cell(row=lastrowm, column=16).value:
                lastrowm += 1
            lastrowm -= 1
            self.process_sheet(wbinput[n], wsout, lastrow+1, lastrowm+1)
        if write_to_file:
            fname, extension = self.output_file.split(".")
            wboutput.save(self.output_file)

if __name__ == "__main__":
    """Testing"""
    output_file="Z:\\Informacion de estudiante en practica\\Archivos para indicadores de mantto\\MTTO CORRECTIVO - DMS\\Anual\\2020S21-29.xlsm"
    input_file="Z:\\Informacion de estudiante en practica\\Archivos para indicadores de mantto\\MTTO CORRECTIVO - DMS\\2020\\SEMANAS\\Semana 21\\Mtto Correctivo Semanal 21-macro.xlsm"
    main(input_file, output_file,True)